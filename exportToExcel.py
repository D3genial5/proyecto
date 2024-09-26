import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys
import os

def crear_directorio_si_no_existe(directorio):
    if not os.path.exists(directorio):
        os.makedirs(directorio)

# Función para leer un único archivo JSON
def leer_json(archivo_json):
    with open(archivo_json, 'r') as file:
        data = json.load(file)
    return data

# La función ahora lee un archivo final si existe, o combina múltiples archivos JSON
def leer_json_multiples(archivo_json_dir):
    cartones = []
    for archivos in os.listdir(archivo_json_dir):
        if archivos.endswith(".json") and "final" not in archivos:  # Evitar leer el archivo final
            with open(os.path.join(archivo_json_dir, archivos), "r") as file:
                data = json.load(file)
                cartones.extend(data)
    return cartones

# Función para exportar los cartones a Excel con celdas centradas
def exportar_a_excel(cartones, archivo_excel):
    filas_excel = []

    # Recorrer todos los cartones y agregar cada uno a una fila diferente
    for carton in cartones:
        fila_carton = []

        # Intercalar los elementos de las tres filas del cartón en una sola fila
        for col_index in range(9):
            fila_carton.append(carton[0][col_index] if carton[0][col_index] is not None else '-')  # Primera fila
            fila_carton.append(carton[1][col_index] if carton[1][col_index] is not None else '-')  # Segunda fila
            fila_carton.append(carton[2][col_index] if carton[2][col_index] is not None else '-')  # Tercera fila

        filas_excel.append(fila_carton)  # Agregar el cartón a la lista de filas

    # Crear un DataFrame y exportarlo a Excel sin index ni header
    df = pd.DataFrame(filas_excel)
    df.to_excel(archivo_excel, index=False, header=False)

    # Cargar el archivo Excel para aplicar formato
    workbook = load_workbook(archivo_excel)
    sheet = workbook.active

    # Aplicar alineación centrada a todas las celdas
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Guardar los cambios en el archivo Excel
    workbook.save(archivo_excel)

# Nueva función para exportar el serial a un archivo .txt
def exportar_serial_a_txt(serial, output_dir):
    archivo_txt = os.path.join(output_dir, f"{serial}_serial.txt")
    with open(archivo_txt, 'w') as file:
        file.write(serial)
    print(f"Serial exportado exitosamente a {archivo_txt}")

# Extraer el serial del nombre del directorio
def extraer_serial_desde_ruta(ruta_json):
    return os.path.basename(ruta_json)  # Asume que el nombre del directorio es el serial

if __name__ == "__main__":
    ruta_json = sys.argv[1]  # Primer argumento: ruta del archivo o directorio JSON
    archivo_excel = sys.argv[2]  # Segundo argumento: ruta del archivo Excel

    # Comprobar si la ruta es un archivo o un directorio
    if os.path.isdir(ruta_json):
        # Es un directorio, leer múltiples archivos JSON
        cartones = leer_json_multiples(ruta_json)
        serial = extraer_serial_desde_ruta(ruta_json)  # Extraer el serial desde la ruta del directorio
    else:
        # Es un archivo único, leer un solo JSON
        cartones = leer_json(ruta_json)
        serial = os.path.splitext(os.path.basename(ruta_json))[0]  # Extraer el serial desde el nombre del archivo

    # Exportar a Excel
    exportar_a_excel(cartones, archivo_excel)

    # Exportar el serial a un archivo de texto
    output_dir = os.path.dirname(archivo_excel)  # Usar el mismo directorio donde se guarda el archivo Excel
    exportar_serial_a_txt(serial, output_dir)

    print(f"Cartones exportados exitosamente a {archivo_excel}")

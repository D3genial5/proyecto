'''import random

def generar_numeros_aleatorios():
    numeros = random.sample(range(1, 91), 15)
    return numeros

cantidadDeVeces = int(input("Igrese la cantidad de veces que quiere que se generen los numeros: "))
nroCarton = cantidadDeVeces
while cantidadDeVeces != 0:
    for i in range(nroCarton):
        numeros_aleatorios = generar_numeros_aleatorios()
        print(f"El carton numero: {i+1}")
        print(numeros_aleatorios)
        cantidadDeVeces -= 1'''

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import random
from uuid import uuid4

def crear_directorio_si_no_existe(directorio):
    if not os.path.exists(directorio):
        os.makedirs(directorio)

# Función para generar un cartón de bingo
def generar_carton():
    carton = [[None for _ in range(9)] for _ in range(3)]  # 3 filas, 9 columnas
    max_por_columna = 2
    rango_limites = [
        (1, 10), (11, 20), (21, 30), (31, 40), (41, 50), 
        (51, 60), (61, 70), (71, 80), (81, 90)
    ]
    
    contador_por_columna = [0] * 9  # Para contar números por columna
    numeros_por_fila = [0] * 3  # Contador de números por fila
    usados = set()  # Para evitar números repetidos
    
    for columna in range(9):
        fila = random.randint(0, 2)  # Seleccionar fila aleatoria
        min_val, max_val = rango_limites[columna]
        numero_aleatorio = None

        while numero_aleatorio is None or numero_aleatorio in usados:
            numero_aleatorio = random.randint(min_val, max_val)
        
        carton[fila][columna] = numero_aleatorio
        usados.add(numero_aleatorio)
        contador_por_columna[columna] += 1
        numeros_por_fila[fila] += 1

    while any(num < 5 for num in numeros_por_fila):
        fila = random.randint(0, 2)
        if numeros_por_fila[fila] >= 5:
            continue

        columna = random.randint(0, 8)
        min_val, max_val = rango_limites[columna]
        numero_aleatorio = None

        if contador_por_columna[columna] < max_por_columna and carton[fila][columna] is None:
            while numero_aleatorio is None or numero_aleatorio in usados:
                numero_aleatorio = random.randint(min_val, max_val)

            carton[fila][columna] = numero_aleatorio
            usados.add(numero_aleatorio)
            contador_por_columna[columna] += 1
            numeros_por_fila[fila] += 1

    for columna in range(9):
        numeros_columna = [carton[fila][columna] for fila in range(3) if carton[fila][columna] is not None]
        numeros_columna.sort()

        index = 0
        for fila in range(3):
            if carton[fila][columna] is not None:
                carton[fila][columna] = numeros_columna[index]
                index += 1
    
    return carton

# Función para generar múltiples cartones de forma secuencial
def generar_cartones_secuencial(cantidad):
    print(f"Comenzando la generación de {cantidad} cartones...")
    cartones = []
    for i in range(cantidad):
        cartones.append(generar_carton())
        if i % 100 == 0:  # Imprimir cada 100 cartones generados
            print(f"{i} cartones generados...")
    print(f"Terminada la generación de {cantidad} cartones.")
    return cartones

# Función para guardar los cartones en archivos JSON
def guardar_cartones(cartones, bloque_size, serial, output_dir):
    bloque_numero = 1
    for i in range(0, len(cartones), bloque_size):
        bloque_cartones = cartones[i:i+bloque_size]
        archivo_json = os.path.join(output_dir, f"{serial}_cartones_bloque_{bloque_numero}.json")
        with open(archivo_json, 'w') as f:
            json.dump(bloque_cartones, f, indent=2)
        print(f"Guardado bloque {bloque_numero} en {archivo_json}")
        bloque_numero += 1

    archivo_final = os.path.join(output_dir, f"{serial}_cartones_final.json")
    with open(archivo_final, 'w') as f:
        json.dump(cartones, f, indent=2)
    print(f"Archivo final de cartones guardado en: {archivo_final}")
    return archivo_final

# Función para exportar el serial a un archivo txt
def exportar_serial_a_txt(serial, output_dir):
    archivo_txt = os.path.join(output_dir, f"{serial}_serial.txt")
    with open(archivo_txt, 'w') as file:
        file.write(serial)
    print(f"Serial exportado exitosamente a {archivo_txt}")

# Función para exportar los cartones a Excel
def exportar_a_excel(cartones, archivo_excel):
    filas_excel = []

    for carton in cartones:
        fila_carton = []
        for col_index in range(9):
            fila_carton.append(carton[0][col_index] if carton[0][col_index] is not None else '-')
            fila_carton.append(carton[1][col_index] if carton[1][col_index] is not None else '-')
            fila_carton.append(carton[2][col_index] if carton[2][col_index] is not None else '-')
        filas_excel.append(fila_carton)

    df = pd.DataFrame(filas_excel)
    df.to_excel(archivo_excel, index=False, header=False)

    workbook = load_workbook(archivo_excel)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(archivo_excel)

if __name__ == "__main__":
    cantidad = int(input("¿Cuántos cartones quieres generar?: "))
    bloque_size = int(input("¿Tamaño del bloque?: "))
    archivo_excel = input("Introduce la ruta completa para el archivo Excel: ")

    serial = str(uuid4())
    serial_root = os.path.join(os.path.dirname(archivo_excel), 'seriales')
    crear_directorio_si_no_existe(serial_root)

    output_dir = os.path.join(serial_root, serial)
    crear_directorio_si_no_existe(output_dir)

    archivo_excel = os.path.join(output_dir, f"{serial}_cartones.xlsx")

    print("Generando cartones de forma secuencial...")
    cartones = generar_cartones_secuencial(cantidad)
    
    print("Guardando cartones en bloques...")
    archivo_final_json = guardar_cartones(cartones, bloque_size, serial, output_dir)

    print("Leyendo y exportando cartones a Excel...")
    with open(archivo_final_json, 'r') as f:
        cartones = json.load(f)

    exportar_a_excel(cartones, archivo_excel)
    exportar_serial_a_txt(serial, output_dir)

    print(f"Cartones exportados exitosamente a {archivo_excel}")